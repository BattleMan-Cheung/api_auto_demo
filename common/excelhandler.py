import json

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from config.setting import config


class ExcelHandler:

    def __init__(self, file):
        self.file = file

    def open_sheet(self, sheet_name) -> Worksheet:
        wb = openpyxl.load_workbook(self.file)
        wb.close()
        return wb[sheet_name]

    def get_title(self, sheet_name):
        sheet = self.open_sheet(sheet_name)
        title = []

        for cell in sheet[1]:
            title.append(cell.value)

        return title

    def read_data(self, sheet_name):
        title = self.get_title(sheet_name)
        sheet = self.open_sheet(sheet_name)

        data = []

        for row in list(sheet)[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(dict(zip(title, row_data)))

        return data

    @staticmethod
    def write_data(file, sheet_name, row, column, data):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row, column).value = data
        wb.save(file)
        wb.close()


if __name__ == '__main__':
    eh = ExcelHandler(config.data_path)
    data = eh.read_data('register')

    print(json.dumps(data))
